from openpyxl.formatting.rule import ColorScaleRule
import numpy as np
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from misc import DBConfig
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter, StrMethodFormatter
from product_and_pricing_tool import PVPricing
from openpyxl import load_workbook
import os
from openpyxl.styles import Font, Alignment
import math
from datetime import date

curr_path = os.path.abspath(__file__)
curr_dir = os.path.dirname(curr_path)
parent_dir = os.path.join(curr_dir, os.pardir)
today_str = date.today().strftime("%Y%m%d")
today_col = date.today().strftime("%Y/%m/%d")


"""
INPUT PARAMETERS
"""
input_param_path = curr_dir + r'\inputs\paper_simulation.xlsx'
df: pd.Series = pd.read_excel(
    io=input_param_path,
    sheet_name='portfolio',
    header=0,
    index_col=0
).squeeze().reset_index().rename(columns={"index": "BloombergCode"})

df_greeks: pd.Series = pd.read_excel(
    io=input_param_path,
    sheet_name='BBG_Greeks',
    header=0,
    index_col=0
).squeeze().reset_index()

df_greeks.insert(0, "AsOfDate", today_col)
output_path0 = curr_dir + fr"\outputs\\{today_str}\\BBG_Greeks_ref_{today_str}.xlsx"
df_greeks.to_excel(output_path0, index=False)

df_pnl: pd.Series = pd.read_excel(
    io=input_param_path,
    sheet_name='PnL_Calc',
    header=0,
    index_col=0
).squeeze().reset_index()

df_pnl.insert(0, "AsOfDate", today_col)
output_path0 = curr_dir + fr"\outputs\\{today_str}\\PnL_Calc_{today_str}.xlsx"
df_pnl.to_excel(output_path0, index=False)

"""
FETCH DATA
"""

col_s = "OPT_UNDL_PX"
col_s_fut = "Undl_fut_px_last"
col_k = "OPT_STRIKE_PX"
col_r = "Cash_rate"
col_tau = "TTM"
col_sigma = "IVOL_MID"
col_q = "EQY_DVD_YLD_EST"
col_sec = "SECURITY_TYP2"
col_mifid = "MIFID_UNDERLYING_DERIVATIVE_TYPE"
col_mid = "PX_MID"
col_position = "Position"
col_pair = "CrncyRatio"

required_cols = [col_s, col_k, col_r, col_tau, col_sigma, col_q]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise KeyError(f"Required columns missing from DataFrame: {missing}\n"
                   f"Existing columns: {list(df.columns)}")


"""
CHECK OPTION
"""
is_option = df.get("SECURITY_TYP2", "").astype(str).str.lower().eq("option")
is_put = df.get("OPT_PUT_CALL", "").astype(str).str.upper().str.startswith("P")
is_call = df.get("OPT_PUT_CALL", "").astype(str).str.upper().str.startswith("C")


def to_num(s):
    return pd.to_numeric(s, errors="coerce")
s = to_num(df[col_s])
s_fut = to_num(df[col_s_fut])
k = to_num(df[col_k])
r = to_num(df[col_r])
tau = to_num(df[col_tau])
sigma = to_num(df[col_sigma])
q = to_num(df[col_q])
position = to_num(df[col_position])
pair = to_num(df[col_pair])

valid_option = (s.gt(0) & k.gt(0) & sigma.gt(0) & tau.gt(0) & r.notna())
valid_put = is_option & is_put & valid_option
valid_call = is_option & is_call & valid_option


df["BS_CALL"] = np.nan
df["BS_PUT"] = np.nan

# OPTION PUT
df.loc[valid_put, "BS_PUT"] = df.loc[valid_put].apply(
    lambda row: PVPricing.bs_put_pricing(
        float(row[col_s]), float(row[col_k]), float(row[col_r]),
        float(row[col_q]) if col_q in df.columns and not pd.isna(row[col_q]) else 0.0,
        float(row[col_tau]), float(row[col_sigma])
    ),
    axis=1
)

# OPTION CALL
df.loc[valid_call, "BS_CALL"] = df.loc[valid_call].apply(
    lambda row: PVPricing.bs_call_pricing(
        float(row[col_s]), float(row[col_k]), float(row[col_r]),
        float(row[col_q]) if col_q in df.columns and not pd.isna(row[col_q]) else 0.0,
        float(row[col_tau]), float(row[col_sigma])
    ),
    axis=1
)

"""
CHECK FUTURE
"""
valid_fut = (s_fut.gt(0) & tau.ge(0) & r.notna())
df["FUT"] = np.nan
df.loc[valid_fut, "FUT"] = df.loc[valid_fut].apply(
    lambda row: float(row[col_s_fut]) * math.exp(
        (float(row[col_r]) - (float(row[col_q]) if col_q in df.columns and pd.notna(row[col_q]) else 0.0))
        * float(row[col_tau])
    ),
    axis=1
)

"""
MERGE OPT and FUT into PV_O_VOL_0
"""

df = df[df[col_tau] >= 0].copy()
col_spot = "PX_LAST"
df["PV_0_VOL_0"] = (
    df[["BS_CALL", "BS_PUT", "FUT"]]
      .bfill(axis=1)
      .iloc[:, 0]
      .fillna(df[col_spot]))
df.insert(0, "AsOfDate", today_col)

output_path1 = curr_dir + fr"\outputs\\{today_str}\\calculated_price_{today_str}.xlsx"
df.to_excel(output_path1, index=False)

#print (df)
"""
SCENARIO BUILDING
"""
spot_shifts = np.linspace(-0.1, 0.1, 21)
vol_shifts  = np.linspace(-0.05, 0.05, 11)

def scenario_prices(row):
    """
    Generate 49 scenario prices (7×7 spot × vol grid) for a single row.
    Spot shift: applied to both futures and options
    Vol shift: applied only to options
    """
    sec_type = str(row[col_sec]).strip().lower()
    mifid = str(row[col_mifid]).strip().upper()
    spot = to_num(row[col_s])
    spot_fut = to_num(row[col_s_fut])
    K = to_num(row[col_k])
    r_ = to_num(row[col_r])
    tau_ = to_num(row[col_tau])
    vol_ = to_num(row[col_sigma])
    q_ = to_num(row[col_q]) if not pd.isna(row[col_q]) else 0.0
    opt_type = str(row.get("OPT_PUT_CALL", "")).strip().upper()
    pv_base = to_num(row["PV_0_VOL_0"])
    position = to_num(row[col_position])
    pair = to_num(row[col_pair])

    # Scenario axis
    spot_shifts = np.linspace(-0.1, 0.1, 21)
    vol_shifts = np.linspace(-0.05, 0.05, 11)

    # VOLI spot shift weight
    if tau_ > 0:
        spot_weight = 0.5 / (np.sqrt(tau_) + np.sqrt(((tau_ * 360) + 30) / 360))
    else:
        spot_weight = 0.0

    scenario_result = []

    for s_shift in spot_shifts:

        # --- SPOT SHIFT ---
        if sec_type != "option":
            shifted_spot = spot_fut * (1 + s_shift)

        else:
            if mifid == "VOLI":
                shifted_spot = spot + s_shift * 100 * spot_weight
            else:
                shifted_spot = spot * (1 + s_shift)

        # --- VOL LOOP ---
        for v_shift in vol_shifts:

            # FUTURE: no vol shift
            if sec_type != "option": #for non-fut, the tau=0, pv=shifted_spot only
                pv = (shifted_spot * math.exp((r_ - q_) * tau_)-pv_base)*position*pair
                scenario_result.append(pv)
                continue

            # OPTION
            if sec_type == "option":
                shifted_vol = vol_ + v_shift
                if shifted_vol <= 0 or tau_ <= 0:
                    intrinsic = max(shifted_spot - K, 0) if opt_type.startswith("C") else max(K - shifted_spot, 0)
                    scenario_result.append((intrinsic - pv_base) * position * pair)
                    continue

                if opt_type.startswith("C"):
                    pv = (PVPricing.bs_call_pricing(
                        shifted_spot, K, r_, q_, tau_, shifted_vol
                    )-pv_base)*position*pair
                else:
                    pv = (PVPricing.bs_put_pricing(
                        shifted_spot, K, r_, q_, tau_, shifted_vol
                    )-pv_base)*position*pair

                scenario_result.append(pv)
                continue

            # OTHER products
            scenario_result.append(spot)

    return scenario_result

# 1. Compute scenario prices
df["SCENARIO_PRICES"] = df.apply(scenario_prices, axis=1)

# 2. Create scenario names
scenario_names = [
    f"PV_S#_{s:.02f}_V#_{v:.02f}"
    for s in spot_shifts
    for v in vol_shifts]

# 3. Build scenario DF
scenario_df = pd.DataFrame(
    df["SCENARIO_PRICES"].tolist(),
    columns=scenario_names)

# 4. Build ID DF (only keep columns that exist)
id_cols = [
    c for c in [
        "BloombergCode",
        "SECURITY_TYP2",
        "OPT_UNDL_TICKER/UNDL_SPOT_TICKER",
        "MIFID_UNDERLYING_ASSET_CLASS",
        "MIFID_UNDERLYING_DERIVATIVE_TYPE"
    ]
    if c in df.columns]

id_df = df[id_cols].reset_index(drop=True)

# 5. Final combined table (ID + scenarios)
pv_scenario_df = pd.concat([id_df, scenario_df], axis=1)
pv_scenario_df.insert(0, "AsOfDate", today_col)

output_path2 = curr_dir + fr"\outputs\\{today_str}\\pv_scenarios_{today_str}.xlsx"
pv_scenario_df.to_excel(output_path2, index=False)


"""
MATRIX BUILDING
"""

def scenario_matrix(prices):
    return pd.DataFrame(
        np.array(prices).reshape(21, 11),
        index=[f"S#{v:+.2%}" for v in spot_shifts],   # vertical (rows)
        columns=[f"V#{v:+.2%}" for v in vol_shifts]   # horizontal (columns)
    )

underlying_col = "OPT_UNDL_TICKER/UNDL_SPOT_TICKER"
pv_by_underlying = pv_scenario_df.groupby(underlying_col).sum()
df["UNDERLYING"] = df[underlying_col]

output_path3 = curr_dir + fr"\outputs\\{today_str}\\sensitivity_tables_{today_str}.xlsx"

with pd.ExcelWriter(output_path3, engine="openpyxl") as writer:
    agg_dict = {}
    # 1. Underlying AGG sheets
    for undl, group in df.groupby("UNDERLYING"):
        agg_prices = np.nansum(np.vstack(group["SCENARIO_PRICES"].to_list()), axis=0)
        agg_matrix = scenario_matrix(agg_prices)
        sheet_name = f"{undl}_AGG"[:31]
        agg_matrix.to_excel(writer, sheet_name=sheet_name)
        agg_dict[undl] = agg_prices

    # 2. TOTAL = sum of all AGGs (easiest and safest)
    total_prices = np.nansum(np.vstack(list(agg_dict.values())), axis=0)
    total_matrix = scenario_matrix(total_prices)
    total_matrix.to_excel(writer, sheet_name="TOTAL_AGG")

    #PRINT COLOR FOR THE WORKBOOK
    workbook = writer.book

    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Skip if sheet has no data matrix
        if max_row < 2 or max_col < 2:
            continue

        # Data range excluding headers
        data_range = f"A2:{chr(64 + max_col)}{max_row}"

        # 2-color scale: low=red → high=green
        color_rule = ColorScaleRule(
            start_type="min", start_color="F8696B",   # heavy red
            end_type="max",   end_color="63BE7B"      # heavy green
        )
        sheet.conditional_formatting.add(data_range, color_rule)




